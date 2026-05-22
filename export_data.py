"""
ARIA Data Export
Reads all ARIA logs and exports them to a single Excel workbook.

Usage:
    python export_data.py

Output:
    logs/aria_export.xlsx  — one sheet per data category
"""

import os
import json
import glob
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print('openpyxl not installed. Run: pip install openpyxl')
    raise

LOGS_DIR   = 'logs'
OUTPUT     = os.path.join(LOGS_DIR, 'aria_export.xlsx')
HEADER_BG  = 'FF1F3A5F'  # dark blue
HEADER_FG  = 'FFFFFFFF'  # white
ALT_ROW_BG = 'FFE8EEF7'  # light blue


# ── Helpers ────────────────────────────────────────────────────────────────────

def _header_style(cell):
    cell.font      = Font(bold=True, color=HEADER_FG, size=10)
    cell.fill      = PatternFill('solid', fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _alt_fill(ws, row_num):
    if row_num % 2 == 0:
        for cell in ws[row_num]:
            cell.fill = PatternFill('solid', fgColor=ALT_ROW_BG)

def _write_sheet(ws, headers, rows):
    """Write headers + rows to a worksheet, auto-size columns."""
    ws.row_dimensions[1].height = 18
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    for r, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)
        _alt_fill(ws, r)

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in ws.iter_rows(min_col=col, max_col=col,
                                 min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, 45)

    ws.freeze_panes = 'A2'


def _load_jsonl(path):
    if not os.path.exists(path):
        return []
    events = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    events.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return events


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_lineage(wb):
    """Agent family tree — every agent that ever lived."""
    ws = wb.create_sheet('Lineage')

    # Collect from retired agents + latest checkpoint
    agents = {}

    for path in sorted(glob.glob(os.path.join(LOGS_DIR, 'retired', '*.json'))):
        try:
            r = json.load(open(path))
            aid = r['agent_id']
            hp  = r.get('hyperparams', {})
            agents[aid] = {
                'agent_id':      aid,
                'status':        'Retired',
                'episodes':      r.get('episodes', ''),
                'total_reward':  round(r.get('total_reward', 0), 1),
                'avg_per_ep':    round(r['total_reward'] / max(r['episodes'], 1), 2)
                                 if r.get('episodes') else '',
                'role':          r.get('role', ''),
                'learning_rate': hp.get('learning_rate', ''),
                'epsilon_decay': hp.get('epsilon_decay', ''),
                'hidden_size':   hp.get('hidden_size', ''),
                'n_layers':      hp.get('n_layers', ''),
                'activation':    hp.get('activation', ''),
                'use_skip':      hp.get('use_skip', ''),
                'timestamp':     r.get('timestamp', '')[:19] if r.get('timestamp') else '',
            }
        except Exception:
            pass

    # Active agents from latest checkpoint
    saves = sorted(glob.glob(os.path.join(LOGS_DIR, 'saves', 'checkpoint_ep*.json')))
    lineage_order = []
    if saves:
        try:
            ckpt = json.load(open(saves[-1]))
            lineage_order = ckpt.get('all_ids_ever', [])
            for aid, am in ckpt.get('agent_meta', {}).items():
                if aid not in agents:
                    agents[aid] = {
                        'agent_id':      aid,
                        'status':        'Active',
                        'episodes':      am.get('episodes', ''),
                        'total_reward':  round(am.get('total_reward', 0), 1),
                        'avg_per_ep':    round(am['total_reward'] / max(am['episodes'], 1), 2)
                                         if am.get('episodes') else '',
                        'role':          '',
                        'learning_rate': round(am.get('learning_rate', 0), 6),
                        'epsilon_decay': round(am.get('epsilon_decay', 0), 4),
                        'hidden_size':   am.get('hidden_size', ''),
                        'n_layers':      am.get('n_layers', ''),
                        'activation':    am.get('activation', ''),
                        'use_skip':      am.get('use_skip', ''),
                        'timestamp':     '',
                    }
        except Exception:
            pass

    # Sort by lineage order where possible
    ordered = [a for a in lineage_order if a in agents]
    rest    = [a for a in agents if a not in ordered]
    all_ids = ordered + rest

    headers = ['Agent ID', 'Status', 'Episodes', 'Total Reward', 'Avg/Episode',
               'Role', 'Learning Rate', 'Epsilon Decay', 'Hidden Size',
               'Layers', 'Activation', 'Skip Conn', 'Retired At']
    rows = []
    for i, aid in enumerate(all_ids, 1):
        a = agents[aid]
        rows.append([
            aid, a['status'], a['episodes'], a['total_reward'], a['avg_per_ep'],
            a['role'], a['learning_rate'], a['epsilon_decay'], a['hidden_size'],
            a['n_layers'], a['activation'], a['use_skip'], a['timestamp'],
        ])

    _write_sheet(ws, headers, rows)
    print(f'  Lineage          : {len(rows)} agents')


def build_checkpoints(wb):
    """Progress snapshot at each autosave point."""
    ws = wb.create_sheet('Progress Checkpoints')

    saves = sorted(glob.glob(os.path.join(LOGS_DIR, 'saves', 'checkpoint_ep*.json')))
    headers = ['Episode', 'Generation', 'Timestamp', 'Agent ID',
               'Epsilon', 'Total Reward', 'Episodes Trained',
               'Learning Rate', 'Activation', 'Layers']
    rows = []
    for path in saves:
        try:
            ckpt = json.load(open(path))
            ep   = ckpt.get('episode', '')
            gen  = ckpt.get('generation', '')
            ts   = ckpt.get('timestamp', '')[:19] if ckpt.get('timestamp') else ''
            for aid, am in ckpt.get('agent_meta', {}).items():
                rows.append([
                    ep, gen, ts, aid,
                    round(am.get('epsilon', 0), 4),
                    round(am.get('total_reward', 0), 1),
                    am.get('episodes', ''),
                    round(am.get('learning_rate', 0), 6),
                    am.get('activation', ''),
                    am.get('n_layers', ''),
                ])
        except Exception:
            pass

    _write_sheet(ws, headers, rows)
    print(f'  Checkpoints      : {len(saves)} saves  ({len(rows)} agent rows)')


def build_help_events(wb):
    """All help requests — granted and denied."""
    ws = wb.create_sheet('Help Events')

    events  = _load_jsonl(os.path.join(LOGS_DIR, 'help.jsonl'))
    headers = ['Timestamp', 'Episode', 'Agent', 'Event', 'Reason',
               'Avg Reward', 'Coord Rate', 'Claude Used',
               'LR Adjusted To', 'Eps Decay Adjusted To',
               'Sub Goal', 'Never Ask Again']
    rows = []
    for e in events:
        if e.get('event') not in ('help_applied', 'help_denied'):
            continue
        applied   = e.get('applied', {})
        lr_to     = applied.get('LEARNING_RATE', {}).get('to', '') if applied else ''
        eps_to    = applied.get('EPSILON_DECAY', {}).get('to', '') if applied else ''
        rows.append([
            e.get('timestamp', '')[:19],
            e.get('episode', ''),
            e.get('agent', ''),
            e.get('event', '').replace('_', ' '),
            e.get('reason', '').replace('_', ' '),
            e.get('avg_reward', ''),
            e.get('coord_rate', ''),
            e.get('claude_used', ''),
            lr_to,
            eps_to,
            e.get('sub_goal', ''),
            e.get('never_ask_again', ''),
        ])

    _write_sheet(ws, headers, rows)
    granted = sum(1 for e in events if e.get('event') == 'help_applied')
    denied  = sum(1 for e in events if e.get('event') == 'help_denied')
    print(f'  Help events      : {granted} granted  {denied} denied')


def build_discovered_goals(wb):
    """Self-discovered goals from internal goal discovery."""
    ws = wb.create_sheet('Discovered Goals')

    events  = _load_jsonl(os.path.join(LOGS_DIR, 'discovered_goals.jsonl'))
    headers = ['Timestamp', 'Episode', 'Agent', 'Condition', 'Bonus', 'Lift']
    rows    = []
    for e in events:
        rows.append([
            e.get('timestamp', '')[:19],
            e.get('episode', ''),
            e.get('agent', ''),
            e.get('condition', ''),
            e.get('bonus', ''),
            e.get('lift', ''),
        ])

    _write_sheet(ws, headers, rows)
    print(f'  Discovered goals : {len(rows)}')


def build_lexicon(wb):
    """Signal usage and symbol crystallisation summary (aggregated — not raw)."""
    ws = wb.create_sheet('Lexicon')

    events = _load_jsonl(os.path.join(LOGS_DIR, 'lexicon.jsonl'))

    # Signal usage aggregated
    signal_stats = defaultdict(lambda: {'uses': 0, 'coord_successes': 0,
                                        'symbol': '', 'senders': set()})
    assignments  = []
    compounds    = []
    sequences    = []

    for e in events:
        ev = e.get('event')
        if ev == 'signal':
            idx = e.get('signal_idx', '')
            signal_stats[idx]['uses']          += 1
            signal_stats[idx]['coord_successes'] += (1 if e.get('coord_achieved') else 0)
            signal_stats[idx]['symbol']          = e.get('symbol', '')
            if e.get('sender'):
                signal_stats[idx]['senders'].add(e['sender'])
        elif ev == 'symbol_assigned':
            assignments.append([
                e.get('timestamp', '')[:19],
                e.get('signal_idx', ''),
                e.get('symbol', ''),
                e.get('use_count_at_assignment', ''),
            ])

    # Sheet: signal summary
    headers = ['Signal Index', 'Symbol', 'Total Uses', 'Coord Successes',
               'Coord Rate', 'Senders']
    rows = []
    for idx, s in sorted(signal_stats.items()):
        coord_rate = (s['coord_successes'] / s['uses']) if s['uses'] > 0 else 0
        rows.append([
            idx,
            s['symbol'],
            s['uses'],
            s['coord_successes'],
            round(coord_rate, 4),
            ', '.join(sorted(s['senders'])),
        ])
    _write_sheet(ws, headers, rows)

    # Second sheet: symbol assignment events
    ws2 = wb.create_sheet('Symbol Assignments')
    _write_sheet(ws2,
                 ['Timestamp', 'Signal Index', 'Symbol', 'Uses at Assignment'],
                 assignments)

    print(f'  Lexicon          : {sum(s["uses"] for s in signal_stats.values()):,} signals  '
          f'{len(assignments)} assignments')


def build_budget(wb):
    """API spending summary."""
    ws = wb.create_sheet('API Budget')

    path = os.path.join(LOGS_DIR, 'budget.json')
    if not os.path.exists(path):
        _write_sheet(ws, ['Note'], [['No API spend recorded yet — add API key to .env']])
        print(f'  API budget       : no spend yet')
        return

    b = json.load(open(path))
    summary_headers = ['Metric', 'Value']
    input_cost  = b['input_tokens']  / 1_000_000 * 0.80
    output_cost = b['output_tokens'] / 1_000_000 * 4.00

    summary_rows = [
        ['Lifetime cost (USD)',    f"${b['lifetime_cost']:.6f}"],
        ['Input tokens',           f"{b['input_tokens']:,}"],
        ['Output tokens',          f"{b['output_tokens']:,}"],
        ['Input cost (USD)',        f"${input_cost:.6f}"],
        ['Output cost (USD)',       f"${output_cost:.6f}"],
        ['Total API calls',         b['total_calls']],
        ['Last updated',            b.get('last_updated', '')[:19]],
        ['', ''],
        ['By call type', ''],
    ]
    for call_type, count in b.get('by_type', {}).items():
        summary_rows.append([f'  {call_type}', count])

    _write_sheet(ws, summary_headers, summary_rows)
    print(f'  API budget       : ${b["lifetime_cost"]:.6f} lifetime  {b["total_calls"]} calls')


def build_retired_detail(wb):
    """Full detail on every retired agent."""
    ws = wb.create_sheet('Retired Agents Detail')

    headers = ['Agent ID', 'Episodes', 'Total Reward', 'Avg Reward/Ep',
               'Epsilon at Retirement', 'Role',
               'Learning Rate', 'Epsilon Decay', 'Intrinsic Beta', 'Episodic Beta',
               'Hidden Size', 'Layers', 'Activation', 'Skip Conn', 'Retired At']
    rows = []
    for path in sorted(glob.glob(os.path.join(LOGS_DIR, 'retired', '*.json'))):
        try:
            r  = json.load(open(path))
            hp = r.get('hyperparams', {})
            ep = r.get('episodes', 0)
            rows.append([
                r.get('agent_id', ''),
                ep,
                round(r.get('total_reward', 0), 1),
                round(r['total_reward'] / max(ep, 1), 2),
                round(r.get('epsilon_at_retirement', 0), 4),
                r.get('role', ''),
                hp.get('learning_rate', ''),
                hp.get('epsilon_decay', ''),
                hp.get('intrinsic_beta', ''),
                hp.get('episodic_beta', ''),
                hp.get('hidden_size', ''),
                hp.get('n_layers', ''),
                hp.get('activation', ''),
                hp.get('use_skip', ''),
                r.get('timestamp', '')[:19] if r.get('timestamp') else '',
            ])
        except Exception:
            pass

    _write_sheet(ws, headers, rows)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print()
    print('  ARIA Data Export')
    print('  ' + '─' * 40)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    build_lineage(wb)
    build_checkpoints(wb)
    build_help_events(wb)
    build_discovered_goals(wb)
    build_lexicon(wb)
    build_budget(wb)
    build_retired_detail(wb)

    os.makedirs(LOGS_DIR, exist_ok=True)
    wb.save(OUTPUT)

    print('  ' + '─' * 40)
    print(f'  Saved to: {OUTPUT}')
    print(f'  Sheets: {", ".join(ws.title for ws in wb.worksheets)}')
    print()


if __name__ == '__main__':
    main()
